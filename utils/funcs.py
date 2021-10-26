#!/usr/bin/env python
# encoding: utf-8
'''
@author: Jerry Lee
@contact: liweizhong@nekteck.com
@software: Yuanrui
@file: funcs.py
@time: 2021/10/14 17:02
@desc:
'''
import os
import socket

import xlrd
import openpyxl


def get_local_ip():
    # 获取本机计算机名称
    hostname = socket.gethostname()

    # 获取本机ip
    ip = socket.gethostbyname(hostname)

    return ip


def read_xls_file(filename):
    """
    从excel文件中读取所有的mac地址
    :param filename:
    :return: testcase 列表
    """
    if os.path.splitext(filename)[-1] == ".xls":
        workbook = xlrd.open_workbook(filename)
        booksheet = workbook.sheet_by_index(0)  # 用索引取第一个sheet

        nrows = booksheet.nrows

        mac_addr_list = []
        for i in range(1, nrows):
            mac_addr = booksheet.cell_value(i, 4)
            mac_addr_list.append(mac_addr)
    else:
        workbook = openpyxl.load_workbook(filename)
        booksheet = workbook.worksheets[0]  # 用索引取第一个sheet

        nrows = booksheet.max_row+1

        mac_addr_list = []
        for i in range(2, nrows):
            mac_addr = booksheet.cell(i, 5).value
            mac_addr_list.append(mac_addr)

    return mac_addr_list
